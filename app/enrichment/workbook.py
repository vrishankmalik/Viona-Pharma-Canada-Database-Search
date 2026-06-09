"""Two-tab enriched workbook builder.

Sheet 1 — "DPD + NOC + Patents":
  One row per DIN (DPD and NOC records joined), sorted ascending by DIN.
  NOC rows whose DIN is blank / "Not Applicable" are excluded entirely.
  Patent block: WIDE format — patent_1_number/filing/grant/expiry through
  patent_M_number/..., where M = max patents held by any DIN in the dataset.
  Labeling columns per DIN (from store). Data protection columns from the
  Register of Innovative Drugs (dp_6yr_no_file_date, pediatric_extension,
  data_protection_ends).

  Columns removed vs old format (Change 2):
    - All *_url columns (record_url, noc_record_url, labeling_pdf_url)
    - All *_page citation columns
    - Old combined patent_numbers / all_patents_detail / earliest_* / latest_* columns
  Columns kept:
    - _drug_code (internal DPD identifier, not a URL)
    - needs_ocr (live OCR-pipeline provenance flag)

Sheet 2 — "Generic Submissions":
  GSUR records filtered to the queried ingredient (same normalisation used
  elsewhere). Standalone — never joined to Sheet 1.

CLI:
  python -m app.enrichment.workbook --q "alpelisib" --field ingredient
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any, Optional

import pandas as pd

from app.config import WORKBOOK_MIN_FILL_RATE

logger = logging.getLogger(__name__)

from app.enrichment.data_protection import (
    _match_data_protection_deterministic,
)
from app.enrichment.store import get_labeling_for_din, get_patents_for_din
from app.models import DrugRecord, SearchResponse

# DIN values that should be excluded from Sheet 1
_EXCLUDED_DIN_VALUES = {"", "not applicable", "n/a", "na", "none"}

# Columns that are NEVER pruned regardless of fill rate.
# This covers every named output column in the Sheet 1 schema.  Only internal
# supplementary DPD fields (_schedule, _last_update) and high-numbered patent tail
# groups (N ≥ 2) are eligible for threshold pruning.
_NEVER_DROP_COLS = frozenset({
    # Identity / provenance
    "din", "_drug_code", "needs_ocr",
    # DPD core
    "brand_name", "company", "ingredient", "strength",
    "dosage_form", "route", "status",
    # Patent summary (patent_1_* is protected separately as the floor group)
    "patent_count",
    # NOC (also guarded by "No NOC record" sentinels, belt-and-suspenders)
    "noc_date", "noc_submission_type", "noc_therapeutic_class",
    # Labeling
    "active_ingredient", "excipients_core", "excipients_coating",
    "preservatives", "pack_size", "pack_style",
    "colour", "shape", "size_mm", "weight", "ph",
    # Data protection (always present even when no record matches)
    "dp_6yr_no_file_date", "pediatric_extension", "data_protection_ends",
})

# Canonical Sheet 1 column order.  Patent N_* columns are inserted between the
# pre-patent and post-patent groups at assembly time (count varies per dataset).
_SHEET1_PRE_PATENT_COLS = (
    "din", "ingredient", "brand_name", "company", "strength", "dosage_form",
    "route", "status", "_drug_code", "_schedule", "_last_update",
    "noc_date", "noc_submission_type", "noc_therapeutic_class",
    "patent_count",
)
_SHEET1_POST_PATENT_COLS = (
    "active_ingredient", "excipients_core", "excipients_coating",
    "preservatives", "pack_size", "pack_style", "colour", "shape", "size_mm",
    "weight", "ph", "needs_ocr",
    "dp_6yr_no_file_date", "pediatric_extension", "data_protection_ends",
)

# Regex to identify patent_N group columns (the four columns per patent slot).
_PATENT_GROUP_RE = re.compile(
    r"^patent_(\d+)_(number|filing_date|grant_date|expiry_date)$"
)


def _is_empty_for_fill(v: Any) -> bool:
    """True if v counts as empty for fill-rate purposes.

    None, NaN, "", and whitespace-only strings are empty.
    Sentinel strings ("No NOC record", "Not in PM", "No", …) are NOT empty —
    they represent a real, meaningful absence and protect the column.
    """
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return str(v).strip() == ""


def _col_fill_rate(series: "pd.Series[Any]", n_rows: int) -> float:
    return int(series.apply(lambda v: not _is_empty_for_fill(v)).sum()) / n_rows


def _prune_sparse_columns(
    df: pd.DataFrame,
    min_fill_rate: float = WORKBOOK_MIN_FILL_RATE,
) -> pd.DataFrame:
    """Drop Sheet 1 columns whose non-empty fill rate is at or below min_fill_rate.

    Patent groups (patent_N_number/filing/grant/expiry) are evaluated and
    dropped together so the wide layout stays aligned.  The report is printed
    to stdout so it appears in CLI and server logs.
    """
    if df.empty:
        return df

    n_rows = len(df)
    cols_before = len(df.columns)
    cols_to_drop: list[str] = []
    report_lines: list[str] = []

    # ── Collect patent groups ────────────────────────────────────────────────
    patent_groups: dict[int, list[str]] = {}
    for col in df.columns:
        m = _PATENT_GROUP_RE.match(col)
        if m:
            patent_groups.setdefault(int(m.group(1)), []).append(col)
    patent_all_cols: set[str] = {c for cols in patent_groups.values() for c in cols}

    # Prune patent tail groups: walk from the HIGHEST group downward, dropping
    # sparse groups until we reach a dense group (or hit the patent_1 floor).
    # patent_1 is NEVER pruned by the threshold — it is the minimum patent slot
    # and its dates represent real data for any DIN that has a patent.
    if patent_groups:
        for n in sorted(patent_groups, reverse=True):
            if n == 1:
                break  # patent_1 is the unconditional floor — stop here
            num_col = f"patent_{n}_number"
            if num_col not in df.columns:
                continue
            fr = _col_fill_rate(df[num_col], n_rows)
            if fr <= min_fill_rate:
                n_filled = round(fr * n_rows)
                cols_to_drop.extend(patent_groups[n])
                report_lines.append(
                    f"  patent_{n} group (4 cols): {n_filled}/{n_rows} = {fr:.1%} fill → dropped"
                )
            else:
                break  # Dense group found — keep this one and everything below it

    # ── Non-patent columns ───────────────────────────────────────────────────
    for col in df.columns:
        if col in patent_all_cols or col in _NEVER_DROP_COLS:
            continue
        fr = _col_fill_rate(df[col], n_rows)
        if fr <= min_fill_rate:
            n_filled = round(fr * n_rows)
            cols_to_drop.append(col)
            report_lines.append(
                f"  {col}: {n_filled}/{n_rows} = {fr:.1%} fill → dropped"
            )

    cols_after = cols_before - len(cols_to_drop)
    print(f"\n=== Workbook column cleanup (min_fill_rate={min_fill_rate:.1%}) ===")
    if report_lines:
        for line in report_lines:
            print(line)
    else:
        print("  (no columns dropped)")
    print(f"  Columns: {cols_before} → {cols_after}")
    print("=" * 52)

    return df.drop(columns=cols_to_drop)

# Supplement submission types to drop from Sheet 1 (SNDS / SANDS)
_SUPPLEMENT_TYPE_RE = re.compile(
    r"\bSNDS\b|\bSANDS\b|Supplement\s+to\s+(a\s+New|an\s+Abbreviated)",
    re.IGNORECASE,
)

# Sentinel dict for DPD DINs that have no matching NOC record
_NO_NOC_RECORD = {
    "noc_date": "No NOC record",
    "noc_submission_type": "No NOC record",
    "noc_therapeutic_class": "No NOC record",
}

_LABELING_FIELDS = (
    "active_ingredient", "excipients_core", "excipients_coating",
    "preservatives", "pack_size", "pack_style",
    "colour", "shape", "size_mm", "weight", "ph",
)


# ── Sheet 1 helpers ───────────────────────────────────────────────────────────

def _is_excluded_din(din: Optional[str]) -> bool:
    return din is None or din.strip().lower() in _EXCLUDED_DIN_VALUES


def _collect_dpd_rows(records: list[DrugRecord]) -> dict[str, dict[str, Any]]:
    """Build DIN-keyed dict from DPD records. record_url excluded per Change 2."""
    out: dict[str, dict[str, Any]] = {}
    for r in records:
        if r.source != "DPD" or _is_excluded_din(r.din):
            continue
        din = r.din.strip()  # type: ignore[union-attr]
        out[din] = {
            "din": din,
            "brand_name": r.brand_name,
            "company": r.company,
            "ingredient": r.ingredient,
            "strength": r.strength,
            "dosage_form": r.dosage_form,
            "route": r.route,
            "status": r.status,
            "_drug_code": r.source_specific.get("drug_code"),
            "_schedule": r.source_specific.get("schedule"),
            "_last_update": r.source_specific.get("last_update_date"),
        }
    return out


def _collect_noc_rows(records: list[DrugRecord]) -> dict[str, dict[str, Any]]:
    """Build DIN-keyed dict from NOC records.

    Supplement rows (SNDS / SANDS) are dropped — only NDS, ANDS, and unknown
    types are included.  noc_record_url excluded per Change 2.
    """
    out: dict[str, dict[str, Any]] = {}
    for r in records:
        if r.source != "NOC" or _is_excluded_din(r.din):
            continue
        sub_type = r.source_specific.get("submission_type") or ""
        if _SUPPLEMENT_TYPE_RE.search(sub_type):
            continue  # drop SNDS / SANDS rows
        din = r.din.strip()  # type: ignore[union-attr]
        out[din] = {
            "noc_date": r.source_specific.get("noc_date"),
            "noc_submission_type": sub_type or None,
            "noc_therapeutic_class": r.source_specific.get("therapeutic_class"),
        }
    return out


def _aggregate_patents_wide(din: str, max_patents: int) -> dict[str, Any]:
    """Return wide patent block: patent_count + max_patents column groups of 4.

    Each group: patent_N_number, patent_N_filing_date, patent_N_grant_date,
    patent_N_expiry_date.  Groups beyond the DIN's actual count are all None.
    max_patents must be >= 1.
    """
    rows = get_patents_for_din(din)
    out: dict[str, Any] = {"patent_count": len(rows)}
    for i in range(1, max_patents + 1):
        if i <= len(rows):
            r = rows[i - 1]
            out[f"patent_{i}_number"] = r["patent_number"]
            out[f"patent_{i}_filing_date"] = r.get("filing_date")
            out[f"patent_{i}_grant_date"] = r.get("grant_date")
            out[f"patent_{i}_expiry_date"] = r.get("expiry_date")
        else:
            out[f"patent_{i}_number"] = None
            out[f"patent_{i}_filing_date"] = None
            out[f"patent_{i}_grant_date"] = None
            out[f"patent_{i}_expiry_date"] = None
    return out


def _get_labeling_cols(din: str) -> dict[str, Any]:
    """Return labeling fields for a DIN.

    *_page citation columns and labeling_pdf_url are excluded per Change 2.
    needs_ocr is kept as a live extraction-provenance flag.
    """
    row = get_labeling_for_din(din)
    out: dict[str, Any] = {}
    for field in _LABELING_FIELDS:
        out[field] = row.get(field) if row else None
        # _page columns intentionally omitted from workbook output
    out["needs_ocr"] = bool(row.get("needs_ocr")) if row else None
    # labeling_pdf_url intentionally omitted from workbook output
    return out


def _get_dp_cols(
    dpd_ingredient: Optional[str],
    dpd_company: Optional[str],
    dp_table: Optional[list[dict]],
) -> dict[str, Any]:
    """Return data protection fields for a DIN, or blanks when dp_table is None."""
    blank = {"dp_6yr_no_file_date": None, "pediatric_extension": None, "data_protection_ends": None}
    if dp_table is None:
        return blank
    matched = _match_data_protection_deterministic(
        dpd_ingredient or "", dpd_company or "", dp_table
    )
    if matched:
        return matched
    return blank


def _col_is_all_empty(series: pd.Series) -> bool:
    """True iff every value in the series is None/NaN/empty string/whitespace.

    Sentinel strings ("No NOC record", "No PM available", "Not in PM", …) are
    NOT empty — they carry real information and prevent the column from dropping.
    """
    for val in series:
        if val is None:
            continue
        try:
            if pd.isna(val):
                continue
        except (TypeError, ValueError):
            pass
        s = str(val).strip()
        if not s or s.lower() in ("none", "nan"):
            continue
        # Has at least one real value (including any sentinel string)
        return False
    return True


def _drop_empty_sheet1_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Drop all-empty columns from Sheet 1 before writing.

    Patent groups (patent_N_number + 3 date cols) are evaluated and dropped as
    a unit: only when ALL four columns in the group are all-empty for all rows.
    This keeps layout aligned when some patents have numbers but missing dates.

    All other columns are dropped individually when all-empty.

    "Empty" means None/NaN/''/whitespace only.  Sentinels like
    'No NOC record', 'No PM available', 'Not in PM' count as REAL data.
    """
    if df.empty:
        return df

    cols_to_drop: list[str] = []
    patent_group_cols: set[str] = set()

    # 1. Patent groups — drop only if the whole group is all-empty
    pat_nums = sorted({
        int(m.group(1))
        for c in df.columns
        if (m := re.match(r"^patent_(\d+)_", c))
    })
    for n in pat_nums:
        grp = [c for c in df.columns if c.startswith(f"patent_{n}_")]
        patent_group_cols.update(grp)
        if all(_col_is_all_empty(df[c]) for c in grp):
            cols_to_drop.extend(grp)

    # 2. Non-patent columns — drop individually if all-empty
    for col in df.columns:
        if col not in patent_group_cols and _col_is_all_empty(df[col]):
            cols_to_drop.append(col)

    if cols_to_drop:
        logger.info(
            "Dropping %d all-empty columns from Sheet 1: %s",
            len(cols_to_drop), cols_to_drop,
        )
        print(f"[workbook] Dropping {len(cols_to_drop)} all-empty Sheet 1 columns: {cols_to_drop}")

    return df.drop(columns=cols_to_drop)


def build_sheet1(
    response: SearchResponse,
    dp_table: Optional[list[dict]] = None,
) -> pd.DataFrame:
    """Build Sheet 1: one row per DIN, DPD + NOC + wide patents + labeling + data protection."""
    all_records = [r for s in response.sources for r in s.records]

    dpd_by_din = _collect_dpd_rows(all_records)
    noc_by_din = _collect_noc_rows(all_records)

    all_dins = sorted(set(dpd_by_din) | set(noc_by_din))
    if not all_dins:
        return pd.DataFrame()

    # Compute M = max patents across all DINs in this result set (at least 1)
    max_patents = max(
        (len(get_patents_for_din(din)) for din in all_dins),
        default=0,
    )
    max_patents = max(max_patents, 1)

    rows = []
    for din in all_dins:
        row: dict[str, Any] = {"din": din}
        row.update(dpd_by_din.get(din, {}))
        # DPD products with no NOC record (or whose NOC submission was filtered out)
        # receive explicit "No NOC record" labels so blanks are unambiguous.
        noc_data = noc_by_din.get(din)
        row.update(noc_data if noc_data is not None else _NO_NOC_RECORD)
        row.update(_aggregate_patents_wide(din, max_patents))
        row.update(_get_labeling_cols(din))
        dpd_rec = dpd_by_din.get(din, {})
        row.update(_get_dp_cols(dpd_rec.get("ingredient"), dpd_rec.get("company"), dp_table))
        rows.append(row)

    df = pd.DataFrame(rows)

    # Apply the canonical column order defined by _SHEET1_PRE/POST_PATENT_COLS.
    # Patent_N_* columns are dynamic (count varies), so they are inserted between
    # the two fixed groups, sorted by patent slot number.
    present = set(df.columns)
    patent_cols = sorted(
        (c for c in present if _PATENT_GROUP_RE.match(c)),
        key=lambda c: (int(_PATENT_GROUP_RE.match(c).group(1)),  # type: ignore[union-attr]
                       ("number", "filing_date", "grant_date", "expiry_date").index(
                           _PATENT_GROUP_RE.match(c).group(2))),  # type: ignore[union-attr]
    )
    ordered = (
        [c for c in _SHEET1_PRE_PATENT_COLS if c in present]
        + patent_cols
        + [c for c in _SHEET1_POST_PATENT_COLS if c in present]
    )
    # Append any remaining columns not yet listed (future-proofing)
    ordered += [c for c in df.columns if c not in set(ordered)]

    df = df[ordered].sort_values("din", kind="stable").reset_index(drop=True)
    return _prune_sparse_columns(df)


# ── Sheet 2 helpers ───────────────────────────────────────────────────────────

def _ingredient_matches(record_ingredient: Optional[str], query: str) -> bool:
    """Return True if query is contained in the record's ingredient string."""
    if not record_ingredient:
        return False
    q = re.sub(r"\s+", " ", query.strip()).lower()
    ing = re.sub(r"\s+", " ", record_ingredient.strip()).lower()
    return q in ing


def build_sheet2(response: SearchResponse) -> pd.DataFrame:
    """Build Sheet 2: Generic Submissions filtered to the queried ingredient."""
    query = response.metadata.query
    rows = []
    for sr in response.sources:
        if sr.source != "GenericSubmissions":
            continue
        for r in sr.records:
            if not _ingredient_matches(r.ingredient, query):
                continue
            rows.append({
                "medicinal_ingredient": r.ingredient,
                "company": r.company,
                "therapeutic_area": r.source_specific.get("therapeutic_area"),
                "year_month_accepted": r.source_specific.get("date_accepted"),
                "status": r.status,
            })

    if not rows:
        return pd.DataFrame(
            columns=["medicinal_ingredient", "company", "therapeutic_area",
                     "year_month_accepted", "status"]
        )
    return pd.DataFrame(rows)


# ── Workbook assembly ─────────────────────────────────────────────────────────

def _style_sheet(worksheet: Any, df: pd.DataFrame) -> None:
    """Apply bold header, freeze row, autofilter, and autosized columns."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=False)

    worksheet.freeze_panes = "A2"

    if not df.empty:
        last_col = get_column_letter(len(df.columns))
        worksheet.auto_filter.ref = f"A1:{last_col}1"

    for i, col in enumerate(df.columns, 1):
        max_val_len = (
            df[col].fillna("").astype(str).str.len().max()
            if not df.empty else 0
        )
        width = min(max(len(str(col)) + 2, int(max_val_len or 0) + 2), 60)
        worksheet.column_dimensions[get_column_letter(i)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)


def _build_status_sheet(
    writer: Any,
    response: SearchResponse,
    source_errors: dict[str, Optional[str]],
) -> None:
    """Write a per-source status warning sheet (used when allow_partial=True)."""
    rows = []
    for src in response.sources:
        rows.append({
            "source": src.source,
            "status": src.status,
            "record_count": src.count,
            "error_message": src.error_message or "",
            "warning": (
                "⚠ DATA MISSING FROM THIS EXPORT"
                if src.status == "error"
                else ""
            ),
        })
    df = pd.DataFrame(
        rows,
        columns=["source", "status", "record_count", "error_message", "warning"],
    )
    sheet_name = "⚠ Source Status"
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    _style_sheet(writer.sheets[sheet_name], df)


def build_workbook_with_data(
    response: SearchResponse,
    source_errors: Optional[dict[str, Optional[str]]] = None,
    dp_table: Optional[list[dict]] = None,
) -> "tuple[bytes, pd.DataFrame, pd.DataFrame]":
    """Assemble the enriched workbook; also return the Sheet 1 and Sheet 2 DataFrames.

    The DataFrames are the authoritative in-memory representation — identical to
    what is written to the XLSX.  Callers that need both (e.g. the dashboard) should
    use this function rather than calling build_workbook() and build_sheet1() separately
    to avoid double-computing the sheets.

    Returns (xlsx_bytes, sheet1_df, sheet2_df).
    """
    buf = io.BytesIO()
    sheet1 = build_sheet1(response, dp_table=dp_table)
    sheet2 = build_sheet2(response)
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet1.to_excel(writer, sheet_name="DPD + NOC + Patents", index=False)
        _style_sheet(writer.sheets["DPD + NOC + Patents"], sheet1)

        sheet2.to_excel(writer, sheet_name="Generic Submissions", index=False)
        _style_sheet(writer.sheets["Generic Submissions"], sheet2)

        if source_errors is not None:
            _build_status_sheet(writer, response, source_errors)

    return buf.getvalue(), sheet1, sheet2


def build_workbook(
    response: SearchResponse,
    source_errors: Optional[dict[str, Optional[str]]] = None,
    dp_table: Optional[list[dict]] = None,
) -> bytes:
    """Assemble the enriched workbook and return XLSX bytes.

    source_errors: when provided (allow_partial=True path), appends a
    '⚠ Source Status' sheet that visibly flags every failed source.
    dp_table: pre-fetched active data protection register rows (from
    fetch_data_protection_table()); None means the three dp_* columns are blank.
    """
    return build_workbook_with_data(response, source_errors, dp_table)[0]


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import asyncio
    import os

    parser = argparse.ArgumentParser(description="Build the two-tab enriched workbook.")
    parser.add_argument("--q", required=True, help="Search query")
    parser.add_argument("--field", default="ingredient",
                        help="ingredient | brand | company | din")
    parser.add_argument("--out", default=None, help="Output file path (.xlsx)")
    _args = parser.parse_args()

    async def _run() -> None:
        from app.sources.dpd import search_dpd
        from app.sources.generic_submissions import search_generic_submissions
        from app.sources.noc import search_noc
        from app.sources.patent_register import search_patent_register
        from app.normalize import normalize_query
        from app.models import SearchMetadata
        from app.enrichment.data_protection import fetch_data_protection_table
        from datetime import datetime, timezone

        canonical, extra_terms = await normalize_query(_args.q, _args.field)
        sources = await asyncio.gather(
            search_dpd(canonical, _args.field, extra_terms),
            search_generic_submissions(canonical, _args.field, extra_terms),
            search_noc(canonical, _args.field, extra_terms),
            search_patent_register(canonical, _args.field, extra_terms),
        )
        response = SearchResponse(
            metadata=SearchMetadata(
                query=_args.q,
                field=_args.field,
                timestamp=datetime.now(timezone.utc).isoformat(),
                normalized_terms=[canonical] + extra_terms,
            ),
            sources=list(sources),
        )

        # Enrich patents
        from app.enrichment.patents import enrich_patents
        all_dins = [
            r.din for s in response.sources for r in s.records
            if r.din and not _is_excluded_din(r.din)
        ]
        if all_dins:
            await enrich_patents(all_dins)

        dp_table = await fetch_data_protection_table()
        xlsx = build_workbook(response, dp_table=dp_table)
        out_path = _args.out or f"enriched_{_args.q.replace(' ', '_')}_{_args.field}.xlsx"
        with open(out_path, "wb") as fh:
            fh.write(xlsx)
        print(f"Wrote {os.path.getsize(out_path):,} bytes → {out_path}")

    asyncio.run(_run())
