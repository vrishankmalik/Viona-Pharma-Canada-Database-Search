"""Tests for app/enrichment/workbook.py.

Three structural tests per the build spec:
  1. NOC rows with DIN = "Not Applicable" / blank are excluded from Sheet 1.
  2. Sheet 1 rows are sorted by DIN ascending.
  3. Sheet 2 (Generic Submissions) is standalone — never joined to Sheet 1.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import pytest

from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult


# ── shared fixture builder ────────────────────────────────────────────────────

def _make_response(
    dpd_records: list[DrugRecord] | None = None,
    noc_records: list[DrugRecord] | None = None,
    gsur_records: list[DrugRecord] | None = None,
    query: str = "alpelisib",
) -> SearchResponse:
    sources = []
    if dpd_records is not None:
        sources.append(SourceResult(source="DPD", status="ok", records=dpd_records))
    if noc_records is not None:
        sources.append(SourceResult(source="NOC", status="ok", records=noc_records))
    if gsur_records is not None:
        sources.append(SourceResult(source="GenericSubmissions", status="ok", records=gsur_records))
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp=datetime.now(timezone.utc).isoformat(),
        ),
        sources=sources,
    )


def _dpd(din: str, brand: str = "BRAND", strength: str = "50 mg") -> DrugRecord:
    return DrugRecord(
        source="DPD",
        din=din,
        brand_name=brand,
        company="Novartis",
        ingredient="alpelisib",
        strength=strength,
        all_ingredients=["alpelisib"],
    )


def _noc(
    din: str,
    brand: str = "BRAND",
    submission_type: str = "NDS",
    submission_class: str = "New Active Substance (NAS)",
    reason_for_supplement: str = "",
    noc_date: str = "2019-05-24",
) -> DrugRecord:
    return DrugRecord(
        source="NOC",
        din=din,
        brand_name=brand,
        company="Novartis",
        ingredient="alpelisib",
        source_specific={
            "noc_date": noc_date,
            "submission_type": submission_type,
            "submission_class": submission_class,
            "reason_for_supplement": reason_for_supplement or None,
        },
    )


def _gsur(ingredient: str = "alpelisib") -> DrugRecord:
    return DrugRecord(
        source="GenericSubmissions",
        ingredient=ingredient,
        company="GenericCo",
        source_specific={"therapeutic_area": "Oncology", "date_accepted": "2022/06"},
    )


# ── Test 1: NOC "Not Applicable" DINs excluded from Sheet 1 ──────────────────

def test_noc_not_applicable_din_excluded():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[
            _noc("02498014"),          # valid — should appear
            _noc("Not Applicable"),    # must be excluded
            _noc(""),                  # blank — must be excluded
            _noc("N/A"),               # N/A — must be excluded
        ],
    )
    df = build_sheet1(response)

    assert not df.empty, "Sheet 1 should have rows"
    dins_in_sheet = set(df["din"].astype(str).str.strip())
    # "Not Applicable", "", "N/A" must not appear
    assert "Not Applicable" not in dins_in_sheet
    assert "" not in dins_in_sheet
    assert "N/A" not in dins_in_sheet
    # Valid DIN from both NOC and DPD should be present
    assert "02498014" in dins_in_sheet


def test_noc_not_applicable_only_gives_empty_sheet1():
    """If all NOC rows have N/A DINs and there are no DPD records, Sheet 1 is empty."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        noc_records=[_noc("Not Applicable"), _noc("N/A")],
    )
    df = build_sheet1(response)
    assert df.empty or len(df) == 0


# ── Test 2: Sheet 1 rows sorted by DIN ascending ─────────────────────────────

def test_sheet1_din_sorted_ascending():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[
            _dpd("02498022"),
            _dpd("00012345"),
            _dpd("02498014"),
        ],
    )
    df = build_sheet1(response)
    dins = list(df["din"].astype(str))
    assert dins == sorted(dins), f"DINs should be sorted ascending, got: {dins}"


def test_sheet1_single_row_still_sorted():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    assert len(df) == 1
    assert df.iloc[0]["din"] == "02498014"


# ── Test 3: Sheet 2 is standalone — no DIN column, only GSUR data ────────────

def test_sheet2_is_standalone_no_din_column():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        gsur_records=[_gsur("alpelisib"), _gsur("alpelisib hydrochloride")],
    )
    df = build_sheet2(response)
    assert "din" not in df.columns, "Sheet 2 must not have a DIN column"
    assert "medicinal_ingredient" in df.columns


def test_sheet2_filtered_to_queried_ingredient():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(
        gsur_records=[
            _gsur("alpelisib"),
            _gsur("alpelisib hydrochloride"),
            _gsur("metformin"),          # unrelated — must NOT appear
            _gsur("metformin hydrochloride"),  # unrelated — must NOT appear
        ],
        query="alpelisib",
    )
    df = build_sheet2(response)
    ings = list(df["medicinal_ingredient"].str.lower())
    assert all("alpelisib" in i for i in ings), (
        f"Sheet 2 should only contain alpelisib submissions, got: {ings}"
    )
    assert not any("metformin" in i for i in ings), "metformin rows must be filtered out"


def test_sheet2_empty_when_no_gsur_source():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet2(response)
    assert len(df) == 0


# ── Test 4: build_workbook produces valid XLSX bytes ─────────────────────────

def test_build_workbook_returns_xlsx():
    from app.enrichment.workbook import build_workbook

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[_noc("02498014")],
        gsur_records=[_gsur()],
    )
    xlsx = build_workbook(response)

    # Should be a valid ZIP/XLSX (starts with PK magic bytes)
    assert xlsx[:2] == b"PK", "XLSX should start with PK (ZIP magic bytes)"

    # Verify both sheets exist
    import zipfile
    with zipfile.ZipFile(io.BytesIO(xlsx)) as zf:
        names = zf.namelist()
    # openpyxl encodes sheet names in xl/worksheets/
    assert any("sheet" in n.lower() for n in names), "No worksheet files found in XLSX"


def test_build_workbook_sheet1_has_patent_columns(tmp_path):
    """patent_count is always present; patent_N groups appear only when data exists."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # Add a patent so the patent_1_* group isn't dropped by the all-empty cleanup
    # patent_count is in _NEVER_DROP so it always survives pruning
    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    assert "patent_count" in df.columns, "patent_count missing"
    # With one seeded patent, patent_1 group is 100% filled — must be kept
    assert "patent_1_number" in df.columns, "patent_1_number missing (1 patent seeded)"
    assert "patent_1_filing_date" in df.columns, "patent_1_filing_date missing"
    assert "patent_1_expiry_date" in df.columns, "patent_1_expiry_date missing"
    # Old merged-string columns must be absent
    assert "patent_numbers" not in df.columns, "stale patent_numbers column present"
    assert "all_patents_detail" not in df.columns, "stale all_patents_detail column present"


def test_build_workbook_sheet1_has_labeling_columns(tmp_path):
    """Labeling columns appear when labeling data is seeded for the DIN."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # Minimal labeling data keeps labeling columns from being pruned.
    store_mod.upsert_labeling("02498014", {
        "active_ingredient": "alpelisib",
        "nonmedicinal_ingredients": "colloidal silicon dioxide, lactose monohydrate",
        "color": "Pink",
        "shape": "Round",
        "ph": "Not stated",
        "needs_ocr": 0,
        "has_unverified": 0,
        "drug_code": 99001,
        "fetched_at": 0,
    })

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    for col in ("nonmedicinal_ingredients", "color", "shape", "ph"):
        assert col in df.columns, f"Expected labeling column '{col}' in Sheet 1"
    # Old columns must be absent
    for old_col in ("excipients_core", "excipients_coating", "preservatives"):
        assert old_col not in df.columns, f"Stale column '{old_col}' must not appear in Sheet 1"


# ── Test 5: patent block aggregation ─────────────────────────────────────────

def test_patent_aggregation(tmp_path):
    """_aggregate_patents_wide correctly produces wide columns for multiple patents."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")
    store_mod.upsert_patent("02498014", "2900000", "2015-01-01", "2020-03-01", "2035-01-01")

    from app.enrichment.workbook import _aggregate_patents_wide
    agg = _aggregate_patents_wide("02498014", 2)

    assert agg["patent_count"] == 2
    # Both patent numbers must appear across the two groups (order is expiry-desc from store)
    patent_numbers = {agg["patent_1_number"], agg["patent_2_number"]}
    assert "2709025" in patent_numbers
    assert "2900000" in patent_numbers
    # Dates must be attached to the correct group
    for i in (1, 2):
        if agg[f"patent_{i}_number"] == "2709025":
            assert agg[f"patent_{i}_filing_date"] == "2008-12-10"
        elif agg[f"patent_{i}_number"] == "2900000":
            assert agg[f"patent_{i}_filing_date"] == "2015-01-01"


# ── Test 6: fail-loud guard — error-vs-no_results distinction ─────────────────

def _make_response_with_noc_error(query: str = "metformin") -> "SearchResponse":
    """Response where NOC has status=error, DPD has results."""
    from app.models import SourceResult
    sources = [
        SourceResult(source="DPD", status="ok", records=[_dpd("02242974")]),
        SourceResult(source="NOC", status="error", error_message="test forced error"),
        SourceResult(source="GenericSubmissions", status="no_results"),
        SourceResult(source="PatentRegister", status="no_results"),
    ]
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp="2026-01-01T00:00:00+00:00",
        ),
        sources=sources,
    )


def test_build_workbook_with_source_errors_adds_warning_sheet():
    """allow_partial=True path: source_errors causes a '⚠ Source Status' sheet to appear."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response_with_noc_error()
    xlsx = build_workbook(response, source_errors={"NOC": "test forced error"})

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    sheet_names = wb.sheetnames
    assert any("Status" in name for name in sheet_names), (
        f"Expected a Source Status warning sheet; got sheets: {sheet_names}"
    )


def test_build_workbook_no_source_errors_no_warning_sheet():
    """Default path (source_errors=None): no warning sheet is added."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")], noc_records=[_noc("02498014")])
    xlsx = build_workbook(response, source_errors=None)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames) == 2, (
        f"Expected exactly 2 sheets with no errors; got: {wb.sheetnames}"
    )


def test_build_workbook_no_results_source_does_not_block():
    """A no_results source (Patent, GSUR) must not add a warning sheet."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")])
    # no_results is NOT an error — source_errors dict should be empty → no sheet
    xlsx = build_workbook(response, source_errors=None)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames) == 2


async def test_export_refuses_with_409_on_source_error(mock_noc, mock_dpd, mock_gsur, mock_patent_register):
    """HTTP 409 when a source is in error and allow_partial is not set."""
    from fastapi.testclient import TestClient
    from app.main import app
    from unittest.mock import AsyncMock, patch
    from app.models import SourceResult

    forced_error = SourceResult(
        source="NOC", status="error", error_message="forced test error"
    )

    with patch("app.main.search_noc", new=AsyncMock(return_value=forced_error)), \
         patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        client = TestClient(app)
        resp = client.get("/api/export?q=metformin&field=ingredient")
        assert resp.status_code == 409, f"Expected 409, got {resp.status_code}: {resp.text}"
        body = resp.json()
        assert "NOC" in body.get("detail", "")
        assert "allow_partial" in body.get("detail", "")


async def test_export_allow_partial_builds_with_warning(mock_noc, mock_dpd, mock_gsur, mock_patent_register):
    """allow_partial=true: builds a workbook with the source-status warning sheet."""
    from fastapi.testclient import TestClient
    import openpyxl
    from app.main import app
    from unittest.mock import AsyncMock, patch
    from app.models import SourceResult

    forced_error = SourceResult(
        source="NOC", status="error", error_message="forced test error"
    )

    with patch("app.main.search_noc", new=AsyncMock(return_value=forced_error)), \
         patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        client = TestClient(app)
        resp = client.get("/api/export?q=metformin&field=ingredient&allow_partial=true")
        assert resp.status_code == 200, f"Expected 200 with allow_partial; got {resp.status_code}"
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert any("Status" in name for name in wb.sheetnames), (
            f"Expected Source Status sheet; got: {wb.sheetnames}"
        )


# ── Test 7: SNDS/SANDS filtering ─────────────────────────────────────────────

def test_snds_rows_included_in_nds_history():
    """SNDS entries are NDS-lineage and must appear in the workbook (not discarded)."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[
            _noc("02498014", submission_type="NDS"),
            _noc("02498022", submission_type="Supplement to a New Drug Submission (SNDS)",
                 submission_class="Other",
                 reason_for_supplement="New formulation"),
        ],
    )
    df = build_sheet1(response)

    # Both DPD DINs must appear (they're DPD products)
    dins = set(df["din"].astype(str))
    assert "02498014" in dins
    assert "02498022" in dins

    # 02498014 should have NDS data; 02498022 should have SNDS data (not "No NOC record")
    row_nds = df[df["din"] == "02498014"].iloc[0]
    row_snds = df[df["din"] == "02498022"].iloc[0]
    assert row_nds["noc_submission_type"] == "NDS"
    assert "SNDS" in str(row_snds["noc_submission_type"]) or "Supplement" in str(row_snds["noc_submission_type"]), (
        f"Expected SNDS data for NDS-lineage SNDS row, got: {row_snds['noc_submission_type']!r}"
    )
    assert row_snds["noc_submission_type"] != "No NOC record", (
        "SNDS entries must not become 'No NOC record' — they are NDS-lineage"
    )
    # reason_for_supplement should be populated for SNDS rows
    assert row_snds["reason_for_supplement"] == "New formulation"


def test_sands_rows_excluded_from_sheet1():
    """SANDS submissions are ANDS-lineage and must be excluded → 'No NOC record'."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498022")],
        noc_records=[
            _noc("02498022", submission_type="Supplement to an Abbreviated New Drug Submission (SANDS)"),
        ],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498022"].iloc[0]
    assert row["noc_submission_type"] == "No NOC record"
    assert row["submission_class"] == "No NOC record"


def test_no_noc_record_din_gets_sentinel():
    """A DPD DIN with no matching NOC record gets 'No NOC record' in all noc_* columns."""
    from app.enrichment.workbook import build_sheet1

    # DPD record for 02498022 but no NOC record for it
    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[_noc("02498014")],  # only 02498014 has NOC
    )
    df = build_sheet1(response)

    row_with_noc = df[df["din"] == "02498014"].iloc[0]
    row_no_noc = df[df["din"] == "02498022"].iloc[0]

    # Row with NOC match should have real data
    assert row_with_noc["noc_submission_type"] == "NDS"
    assert row_with_noc["noc_date"] == "2019-05-24"

    # Row without NOC match should have sentinel in every noc_* column
    for col in ("noc_date", "reason_for_supplement", "submission_class",
                "noc_submission_type", "noc_therapeutic_class"):
        assert row_no_noc[col] == "No NOC record", (
            f"Expected 'No NOC record' in column '{col}', got: {row_no_noc[col]!r}"
        )


def test_noc_only_column_values_after_filtering():
    """After collection: NDS, ANDS, SNDS are valid; SANDS must never appear."""
    import re
    from app.enrichment.workbook import build_sheet1

    _SANDS_RE = re.compile(
        r"\bSANDS\b|Supplement\s+to\s+an\s+Abbreviated", re.IGNORECASE
    )
    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022"), _dpd("02498030")],
        noc_records=[
            _noc("02498014", submission_type="NDS"),
            _noc("02498022", submission_type="Abbreviated New Drug Submission (ANDS)"),
            _noc("02498030", submission_type="Supplement to a New Drug Submission (SNDS)",
                 submission_class="Other", reason_for_supplement="Extended indication"),
        ],
    )
    df = build_sheet1(response)

    for val in df["noc_submission_type"].dropna():
        for part in str(val).split("\n"):
            part = part.strip()
            if part and part != "No NOC record":
                assert not _SANDS_RE.search(part), (
                    f"SANDS must not appear in noc_submission_type, got: {part!r}"
                )

    # SNDS DIN should have real SNDS data (not "No NOC record")
    row_snds = df[df["din"] == "02498030"].iloc[0]
    assert row_snds["noc_submission_type"] != "No NOC record", (
        "SNDS entry (NDS-lineage) must not become 'No NOC record'"
    )


# ── Test 8: No PM available sentinel ─────────────────────────────────────────

def test_no_pm_available_constant_exported():
    """NO_PM_AVAILABLE constant must be importable and distinct from NOT_IN_PM."""
    from app.enrichment.labeling import NO_PM_AVAILABLE, NOT_IN_PM
    assert NO_PM_AVAILABLE != NOT_IN_PM
    assert NO_PM_AVAILABLE == "No PM available"


# ── Test 9: Change 2 — all-empty column cleanup ──────────────────────────────

def test_empty_patent_groups_dropped(tmp_path):
    """When no DIN has patents, all patent_N_* groups must be dropped.

    patent_count stays (0 is a real integer value, not empty); the four-column
    patent_1_* group is dropped because every cell is None.
    """
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # Do NOT add any patents → all patent_N_* cols will be None

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    # The four-column patent group must have been dropped
    assert "patent_1_number" not in df.columns, (
        "patent_1_number must be dropped when no DINs have patents"
    )
    assert "patent_1_filing_date" not in df.columns
    assert "patent_1_grant_date" not in df.columns
    assert "patent_1_expiry_date" not in df.columns
    # patent_count is 0 (integer) → not empty → must stay
    assert "patent_count" in df.columns, "patent_count must stay even when all counts are 0"


def test_patent_group_kept_when_one_din_has_patent(tmp_path):
    """When even one DIN has a patent, the patent_N_* group must be kept
    for ALL rows (trailing DINs get None in those cols, which is correct)."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")
    # 02498022 has no patents but the group must still exist because 02498014 does

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    assert "patent_1_number" in df.columns, "patent_1_number must be kept (02498014 has a patent)"
    row_with = df[df["din"] == "02498014"].iloc[0]
    row_without = df[df["din"] == "02498022"].iloc[0]
    assert row_with["patent_1_number"] == "2709025"
    assert row_without["patent_1_number"] is None or str(row_without["patent_1_number"]) in (
        "None", "nan", ""
    )


def test_sentinel_values_prevent_column_drop(tmp_path):
    """Columns containing only sentinel strings ('No NOC record', 'No PM available')
    must NOT be dropped — sentinels convey real information."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_sheet1

    # DPD record only → noc_* columns get "No NOC record" sentinel
    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    # All rows have "No NOC record" in noc_* columns — must NOT be dropped
    for col in ("noc_date", "reason_for_supplement", "submission_class",
                "noc_submission_type"):
        assert col in df.columns, (
            f"{col} must be kept even when all values are the 'No NOC record' sentinel"
        )


def test_single_nonempty_row_prevents_column_drop(tmp_path):
    """A column with even one non-empty value must not be dropped,
    even if all other rows are None."""
    import time
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    # Give one DIN a labeling color; other DIN has nothing
    store_mod.upsert_labeling("02498014", {
        "color": "blue", "needs_ocr": 0, "has_unverified": 0,
        "drug_code": 99001, "fetched_at": time.time(),
    })
    # 02498022 has no labeling → color=None for that row

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    assert "color" in df.columns, (
        "color must be kept: 02498014 has a non-empty value even though 02498022 is None"
    )


def test_export_never_produces_old_sheet_names():
    """The /api/export endpoint must never produce the old multi-sheet format."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[_noc("02498014")],
        gsur_records=[_gsur()],
    )
    xlsx = build_workbook(response)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    old_names = {
        "Drug Product Database (DPD)",
        "Generic Submissions Under Review",
        "Notice of Compliance (NOC)",
        "Patent Register",
        "Combined",
        "By Combination",
        "Search Metadata",
    }
    for sheet_name in wb.sheetnames:
        assert sheet_name not in old_names, (
            f"Old sheet name '{sheet_name}' must not appear in the new two-sheet workbook"
        )
    assert set(wb.sheetnames) == {"DPD + NOC + Patents", "Generic Submissions"}, (
        f"Expected exactly the two new sheet names; got: {wb.sheetnames}"
    )


# ── Test 9: threshold-based column pruning ────────────────────────────────────

def test_sparse_patent_group_pruned():
    """patent_7 group with 1/200 fill is dropped; patent_1 is the unconditional floor."""
    import pandas as pd
    from app.enrichment.workbook import _prune_sparse_columns

    n = 200
    # Build a frame with:
    #   patent_1: only 1/200 filled (0.5%) — below threshold BUT must NEVER be dropped
    #   patent_7: 1/200 filled (0.5%) — above patent_1, must be dropped as a tail group
    # Also a real column at 50% fill, and a stray non-patent column at 0.5% fill.
    data: dict[str, list] = {
        "din": [f"{i:08d}" for i in range(n)],
        "patent_count": [1] + [0] * (n - 1),
        "patent_1_number": (["2709025"] + [None] * (n - 1)),
        "patent_1_filing_date": (["2008-12-10"] + [None] * (n - 1)),
        "patent_1_grant_date": (["2014-08-26"] + [None] * (n - 1)),
        "patent_1_expiry_date": (["2028-12-10"] + [None] * (n - 1)),
        "patent_7_number": (["3000001"] + [None] * (n - 1)),
        "patent_7_filing_date": (["2020-01-01"] + [None] * (n - 1)),
        "patent_7_grant_date": ([None] * n),
        "patent_7_expiry_date": ([None] * n),
        # A schema column at 50% fill — must survive (also in _NEVER_DROP_COLS)
        "noc_date": (["2019-05-24"] * (n // 2) + [None] * (n // 2)),
        # dp_6yr_no_file_date is a protected schema column — must survive even at low fill
        "dp_6yr_no_file_date": ([None] * (n - 1) + ["2030-01-01"]),
        # _schedule is a supplementary internal field — eligible for pruning
        "_schedule": ([None] * (n - 1) + ["Rx"]),
    }
    df = pd.DataFrame(data)

    pruned = _prune_sparse_columns(df, min_fill_rate=0.02)

    # patent_7 group must be gone entirely (tail group, N > 1, below threshold)
    for col in ("patent_7_number", "patent_7_filing_date",
                "patent_7_grant_date", "patent_7_expiry_date"):
        assert col not in pruned.columns, f"{col} should have been pruned"

    # patent_1 group must ALWAYS survive — it is the unconditional floor
    # even at 0.5% fill (1/200 rows)
    for col in ("patent_1_number", "patent_1_filing_date",
                "patent_1_grant_date", "patent_1_expiry_date"):
        assert col in pruned.columns, (
            f"{col} must survive: patent_1 is the unconditional floor "
            f"and must never be pruned by the fill-rate threshold"
        )

    # 50%-filled schema column must survive
    assert "noc_date" in pruned.columns, "noc_date (50% fill) must not be pruned"

    # dp_6yr_no_file_date is a protected schema column — must survive regardless of fill
    assert "dp_6yr_no_file_date" in pruned.columns, (
        "dp_6yr_no_file_date must never be pruned (protected schema column)"
    )

    # _schedule is supplementary — eligible for pruning at 0.5% fill
    assert "_schedule" not in pruned.columns, (
        "_schedule (1/200 fill) is a supplementary field and must be pruned"
    )

    # din and patent_count are never dropped
    assert "din" in pruned.columns
    assert "patent_count" in pruned.columns


def test_patent_tail_prune_stops_at_dense_group():
    """Pruning walks from highest N downward and stops at the first dense group.

    If patent_5 through patent_9 are sparse but patent_4 is dense,
    groups 5-9 are dropped and 1-4 are kept — even if 2 and 3 are sparse.
    """
    import pandas as pd
    from app.enrichment.workbook import _prune_sparse_columns

    n = 100
    data: dict[str, list] = {
        "din": [f"{i:08d}" for i in range(n)],
        "patent_count": [4] * 30 + [0] * 70,
        # Groups 1-4: well-filled (30% of rows)
        **{f"patent_{g}_number": (["PAT"] * 30 + [None] * 70) for g in range(1, 5)},
        **{f"patent_{g}_filing_date": (["2010-01-01"] * 30 + [None] * 70) for g in range(1, 5)},
        **{f"patent_{g}_grant_date": (["2015-01-01"] * 30 + [None] * 70) for g in range(1, 5)},
        **{f"patent_{g}_expiry_date": (["2030-01-01"] * 30 + [None] * 70) for g in range(1, 5)},
        # Groups 5-7: sparse (1 row each)
        **{f"patent_{g}_number": ([None] * (g - 1) + ["PAT"] + [None] * (n - g)) for g in range(5, 8)},
        **{f"patent_{g}_filing_date": ([None] * n) for g in range(5, 8)},
        **{f"patent_{g}_grant_date": ([None] * n) for g in range(5, 8)},
        **{f"patent_{g}_expiry_date": ([None] * n) for g in range(5, 8)},
    }
    df = pd.DataFrame(data)
    pruned = _prune_sparse_columns(df, min_fill_rate=0.02)

    # Groups 5-7 must be dropped (sparse tail, N > 1)
    for g in range(5, 8):
        assert f"patent_{g}_number" not in pruned.columns, f"patent_{g} tail must be pruned"

    # Groups 1-4 must survive (dense, pruning stopped before them)
    for g in range(1, 5):
        for suffix in ("number", "filing_date", "grant_date", "expiry_date"):
            assert f"patent_{g}_{suffix}" in pruned.columns, (
                f"patent_{g}_{suffix} must survive (pruning stopped at patent_4)"
            )


def test_sentinel_values_protect_column():
    """Columns filled only with sentinels like 'No NOC record' must NOT be pruned."""
    import pandas as pd
    from app.enrichment.workbook import _prune_sparse_columns

    n = 100
    data: dict[str, list] = {
        "din": [f"{i:08d}" for i in range(n)],
        "patent_count": [0] * n,
        # Every row has "No NOC record" — this is a sentinel, not empty, so must survive
        "noc_submission_type": ["No NOC record"] * n,
        # Column with real value in 60% of rows
        "brand_name": (["PIQRAY"] * 60 + [None] * 40),
    }
    df = pd.DataFrame(data)
    pruned = _prune_sparse_columns(df, min_fill_rate=0.02)

    assert "noc_submission_type" in pruned.columns, (
        "'No NOC record' sentinel must count as non-empty and protect the column"
    )
    assert "brand_name" in pruned.columns


# ── Test 10: new NOC columns — submission_class and reason_for_supplement ─────

def test_new_noc_columns_present_for_nds_record():
    """A plain NDS DIN has submission_class populated and reason_for_supplement blank."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[
            _noc("02498014", submission_type="NDS",
                 submission_class="New Active Substance (NAS)"),
        ],
    )
    df = build_sheet1(response)
    assert "submission_class" in df.columns, "submission_class column missing"
    assert "reason_for_supplement" in df.columns, "reason_for_supplement column missing"
    row = df[df["din"] == "02498014"].iloc[0]
    assert row["submission_class"] == "New Active Substance (NAS)"
    # NDS is not a supplement — reason_for_supplement should be blank/None, not "No NOC record"
    assert row["reason_for_supplement"] is None or str(row["reason_for_supplement"]).strip() in ("", "None")


def test_new_noc_columns_sentinel_for_no_noc_din():
    """A DIN with no NOC record gets 'No NOC record' in all five noc_* columns."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[_noc("02498014")],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498022"].iloc[0]
    for col in ("noc_date", "reason_for_supplement", "submission_class",
                "noc_submission_type", "noc_therapeutic_class"):
        assert row[col] == "No NOC record", (
            f"Column '{col}' should be 'No NOC record' for DIN with no NOC, got: {row[col]!r}"
        )


def test_ands_din_shows_single_entry_no_multi_history():
    """An ANDS/generic DIN shows its ANDS data as a single entry (no multi-history)."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[
            _noc("02498014", submission_type="Abbreviated New Drug Submission (ANDS)",
                 submission_class="Other"),
        ],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498014"].iloc[0]
    assert row["noc_submission_type"] == "Abbreviated New Drug Submission (ANDS)"
    assert row["submission_class"] == "Other"
    # Single entry — no newline separator
    assert "\n" not in str(row["noc_date"])
    assert "\n" not in str(row["noc_submission_type"])


def test_multi_entry_nds_history_aligned():
    """A DIN appearing in two NDS NOC entries gets both joined, positionally aligned."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[
            _noc("02498014", submission_type="NDS",
                 submission_class="New Active Substance (NAS)", noc_date="2014-11-12"),
            _noc("02498014", submission_type="NDS",
                 submission_class="Admin", noc_date="2020-01-29"),
        ],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498014"].iloc[0]

    # Both entries joined with newline
    dates = str(row["noc_date"]).split("\n")
    classes = str(row["submission_class"]).split("\n")
    reasons = str(row["reason_for_supplement"]).split("\n")
    sub_types = str(row["noc_submission_type"]).split("\n")

    assert len(dates) == 2, f"Expected 2 date entries, got: {dates}"
    assert len(classes) == len(dates), "submission_class and noc_date must have equal entry count"
    assert len(reasons) == len(dates), "reason_for_supplement and noc_date must have equal entry count"
    assert len(sub_types) == len(dates), "noc_submission_type and noc_date must have equal entry count"

    # Chronological ascending order
    assert dates[0] == "2014-11-12" and dates[1] == "2020-01-29", (
        f"Expected dates in ascending order, got: {dates}"
    )
    assert classes[0] == "New Active Substance (NAS)" and classes[1] == "Admin"


def test_nds_plus_snds_history_aligned():
    """A DIN with one NDS entry and one SNDS supplement gets both in aligned history."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[
            _noc("02498014", submission_type="NDS",
                 submission_class="New Active Substance (NAS)", noc_date="2010-06-15"),
            _noc("02498014",
                 submission_type="Supplement to a New Drug Submission (SNDS)",
                 submission_class="Other",
                 reason_for_supplement="New formulation for subcutaneous administration",
                 noc_date="2013-02-21"),
        ],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498014"].iloc[0]

    dates = str(row["noc_date"]).split("\n")
    reasons = str(row["reason_for_supplement"]).split("\n")
    classes = str(row["submission_class"]).split("\n")
    sub_types = str(row["noc_submission_type"]).split("\n")

    assert len(dates) == 2
    assert all(len(lst) == 2 for lst in (reasons, classes, sub_types)), (
        "All joined NOC columns must have identical entry count (2)"
    )
    # NDS entry (older): no supplement reason → empty string placeholder
    assert dates[0] == "2010-06-15"
    assert reasons[0] == "", (
        f"NDS entry should have empty reason_for_supplement placeholder, got: {reasons[0]!r}"
    )
    # SNDS entry (newer): supplement reason populated
    assert dates[1] == "2013-02-21"
    assert reasons[1] == "New formulation for subcutaneous administration"
    assert classes[1] == "Other"


def test_ands_excluded_from_nds_lineage_history():
    """A DIN with NDS+SNDS+ANDS entries: ANDS excluded, NDS+SNDS aligned."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[
            _noc("02498014", submission_type="NDS",
                 submission_class="New Active Substance (NAS)", noc_date="2014-01-01"),
            _noc("02498014",
                 submission_type="Supplement to a New Drug Submission (SNDS)",
                 submission_class="Other",
                 reason_for_supplement="Extended indication",
                 noc_date="2016-06-01"),
            _noc("02498014",
                 submission_type="Abbreviated New Drug Submission (ANDS)",
                 submission_class="Other", noc_date="2022-03-10"),
        ],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498014"].iloc[0]

    dates = str(row["noc_date"]).split("\n")
    sub_types = str(row["noc_submission_type"]).split("\n")

    # ANDS entry must be excluded — only NDS + SNDS (2 entries)
    assert len(dates) == 2, f"Expected 2 NDS-lineage entries (ANDS excluded), got: {dates}"
    assert not any("ANDS" in s for s in sub_types), (
        f"ANDS must be excluded from NDS-lineage history: {sub_types}"
    )
    # Chronological order
    assert dates == ["2014-01-01", "2016-06-01"]


def test_alignment_stress_all_noc_columns_equal_entry_count():
    """For every DIN with multi-entry NOC history, all joined columns have identical entry count."""
    from app.enrichment.workbook import build_sheet1

    # Four DINs with varying history lengths
    response = _make_response(
        dpd_records=[_dpd("00000001"), _dpd("00000002"), _dpd("00000003")],
        noc_records=[
            # DIN 00000001: 3 NDS entries
            _noc("00000001", submission_type="NDS", submission_class="NAS", noc_date="2010-01-01"),
            _noc("00000001", submission_type="NDS", submission_class="Admin", noc_date="2015-01-01"),
            _noc("00000001",
                 submission_type="Supplement to a New Drug Submission (SNDS)",
                 submission_class="Other",
                 reason_for_supplement="Expanded indication",
                 noc_date="2018-01-01"),
            # DIN 00000002: 1 NDS entry
            _noc("00000002", submission_type="NDS", submission_class="NAS", noc_date="2012-06-01"),
            # DIN 00000003: ANDS only (single entry)
            _noc("00000003",
                 submission_type="Abbreviated New Drug Submission (ANDS)",
                 submission_class="Other", noc_date="2020-01-01"),
        ],
    )
    df = build_sheet1(response)

    noc_join_cols = ["noc_date", "reason_for_supplement", "submission_class", "noc_submission_type"]
    for _, row in df.iterrows():
        din = row["din"]
        counts = []
        for col in noc_join_cols:
            if col not in df.columns:
                continue
            val = str(row[col]) if row[col] is not None else ""
            if val in ("No NOC record", "None", ""):
                counts.append(1)  # sentinel = single entry
            else:
                counts.append(len(val.split("\n")))
        assert len(set(counts)) <= 1, (
            f"DIN {din}: joined NOC columns have unequal entry counts: "
            + str({c: str(row[c])[:40] for c in noc_join_cols if c in df.columns})
        )


# ═══════════════════════════════════════════════════════════════════════════════
# Multi-product side-by-side workbook tests
# ═══════════════════════════════════════════════════════════════════════════════

def _make_response_for(ingredient: str, dins: list[str], query: str = None) -> "SearchResponse":
    """Build a minimal SearchResponse for a named ingredient with given DINs."""
    query = query or ingredient
    dpd = [
        DrugRecord(
            source="DPD",
            din=din,
            brand_name=f"BRAND-{din}",
            company="TestCo",
            ingredient=ingredient,
            strength="50 mg",
            all_ingredients=[ingredient],
        )
        for din in dins
    ]
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp=datetime.now(timezone.utc).isoformat(),
        ),
        sources=[SourceResult(source="DPD", status="ok", records=dpd)],
    )


class TestMultiProductWorkbook:
    """Tests for build_workbook_multiproduct (side-by-side, color-coded)."""

    def _load_ws(self, xlsx_bytes: bytes, sheet_name: str):
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        return wb[sheet_name]

    def _read_row(self, ws, row: int) -> list:
        return [cell.value for cell in ws[row]]

    def test_single_product_degenerate_valid_xlsx(self):
        """Single product: output is valid XLSX with two tabs."""
        from app.enrichment.workbook import build_workbook_multiproduct
        response = _make_response_for("alpelisib", ["02498014", "02498022"])
        xlsx, s1, s2 = build_workbook_multiproduct([("alpelisib", response)])
        assert xlsx[:2] == b"PK", "must be a valid ZIP/XLSX"
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        assert "DPD + NOC + Patents" in wb.sheetnames
        assert "Generic Submissions" in wb.sheetnames

    def test_key_row_is_row1(self):
        """Row 1 starts with 'PRODUCT KEY:' and has one entry per product."""
        from app.enrichment.workbook import build_workbook_multiproduct
        r1 = _make_response_for("alpelisib", ["02498014"])
        r2 = _make_response_for("apremilast", ["02368975", "02369000"])
        xlsx, _, _ = build_workbook_multiproduct([("alpelisib", r1), ("apremilast", r2)])
        ws = self._load_ws(xlsx, "DPD + NOC + Patents")
        row1 = self._read_row(ws, 1)
        assert row1[0] == "PRODUCT KEY:", f"A1 must be 'PRODUCT KEY:', got {row1[0]!r}"
        assert "alpelisib" in row1, "Product 1 name must be in key row"
        assert "apremilast" in row1, "Product 2 name must be in key row"

    def test_banner_row_is_row2(self):
        """Row 2 contains product names (banner) above each block."""
        from app.enrichment.workbook import build_workbook_multiproduct
        r1 = _make_response_for("alpelisib", ["02498014"])
        r2 = _make_response_for("apremilast", ["02368975"])
        xlsx, _, _ = build_workbook_multiproduct([("alpelisib", r1), ("apremilast", r2)])
        ws = self._load_ws(xlsx, "DPD + NOC + Patents")
        # Banner text includes the product name (uppercased) in the merged cell
        row2_vals = [str(v or "").upper() for v in self._read_row(ws, 2)]
        assert any("ALPELISIB" in v for v in row2_vals), "alpelisib banner missing from row 2"
        assert any("APREMILAST" in v for v in row2_vals), "apremilast banner missing from row 2"

    def test_data_starts_at_row4(self):
        """Row 3 has column headers; actual DIN data begins at row 4."""
        from app.enrichment.workbook import build_workbook_multiproduct
        response = _make_response_for("alpelisib", ["02498014", "02498022"])
        xlsx, _, _ = build_workbook_multiproduct([("alpelisib", response)])
        ws = self._load_ws(xlsx, "DPD + NOC + Patents")
        row3 = [str(v).lower() for v in self._read_row(ws, 3) if v is not None]
        assert "din" in row3, f"'din' header must be in row 3, got: {row3}"
        # Row 4 should contain a DIN value, not a header name
        row4 = [str(v).lower() for v in self._read_row(ws, 4) if v is not None]
        assert row4, "Row 4 must contain data"
        assert "din" not in row4, "Row 4 must be data, not headers"

    def test_no_cross_contamination_two_products(self):
        """Each block contains only its own product's DINs — zero bleed-over."""
        from app.enrichment.workbook import build_workbook_multiproduct
        # Use non-overlapping DINs
        abrocitinib_dins = ["02517019", "02517027", "02517035"]
        apremilast_dins  = [f"0236{i:04d}" for i in range(22)]  # 22 DINs
        r1 = _make_response_for("abrocitinib", abrocitinib_dins)
        r2 = _make_response_for("apremilast",  apremilast_dins)
        xlsx, s1, s2 = build_workbook_multiproduct([("abrocitinib", r1), ("apremilast", r2)])

        # The flat combined DataFrame has a 'product' column
        assert "product" in s1.columns, "combined sheet1 must have 'product' column"
        s1_abro  = s1[s1["product"] == "abrocitinib"]
        s1_aprem = s1[s1["product"] == "apremilast"]

        abro_dins_found  = set(s1_abro["din"].dropna().tolist())
        aprem_dins_found = set(s1_aprem["din"].dropna().tolist())

        # No DIN should appear in both blocks
        overlap = abro_dins_found & aprem_dins_found
        assert not overlap, f"Cross-contamination detected — DINs in both blocks: {overlap}"

        # Each block should contain only its own DINs
        assert abro_dins_found <= set(abrocitinib_dins), "Abrocitinib block has foreign DINs"
        assert aprem_dins_found <= set(apremilast_dins), "Apremilast block has foreign DINs"

    def test_ragged_heights_handled(self):
        """Different DIN counts per product don't corrupt the XLSX."""
        from app.enrichment.workbook import build_workbook_multiproduct
        r1 = _make_response_for("abrocitinib", ["02517019", "02517027", "02517035"])
        r2 = _make_response_for("apremilast",  [f"0236{i:04d}" for i in range(22)])
        xlsx, s1, _ = build_workbook_multiproduct([("abrocitinib", r1), ("apremilast", r2)])
        # Both products' DINs present
        assert len(s1[s1["product"] == "abrocitinib"]) == 3
        assert len(s1[s1["product"] == "apremilast"])  == 22

    def test_single_product_same_data_as_build_workbook_with_data(self):
        """Single-product multiproduct workbook data matches the classic path."""
        from app.enrichment.workbook import build_sheet1, build_workbook_multiproduct
        response = _make_response(
            dpd_records=[_dpd("02498014"), _dpd("02498022")],
            noc_records=[_noc("02498014"), _noc("02498022")],
        )
        _, s1_multi, _ = build_workbook_multiproduct([("alpelisib", response)])
        s1_single = build_sheet1(response)

        # Combined df has extra 'product' column — drop it for comparison
        s1_multi_cmp = s1_multi.drop(columns=["product"]).reset_index(drop=True)
        s1_single_cmp = s1_single.reset_index(drop=True)

        assert list(s1_multi_cmp.columns) == list(s1_single_cmp.columns), (
            "Single-product multi path must produce same columns as build_sheet1"
        )
        assert len(s1_multi_cmp) == len(s1_single_cmp), (
            "Single-product multi path must produce same row count"
        )
        import math
        for col in s1_single_cmp.columns:
            for r_idx, (v_multi, v_single) in enumerate(
                zip(s1_multi_cmp[col], s1_single_cmp[col])
            ):
                # NaN != NaN in Python; treat both-NaN as equal
                both_nan = (
                    (v_multi is None or (isinstance(v_multi, float) and math.isnan(v_multi)))
                    and
                    (v_single is None or (isinstance(v_single, float) and math.isnan(v_single)))
                )
                if not both_nan:
                    assert v_multi == v_single, (
                        f"Column {col!r} row {r_idx}: multi={v_multi!r} vs single={v_single!r}"
                    )

    def test_both_tabs_present_and_color_matched(self):
        """Both tabs have the same products with matching colors."""
        from app.enrichment.workbook import build_workbook_multiproduct
        r1 = _make_response_for("alpelisib", ["02498014"])
        r2 = _make_response_for("apremilast", ["02368975"])
        xlsx, _, _ = build_workbook_multiproduct([("alpelisib", r1), ("apremilast", r2)])
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        assert "DPD + NOC + Patents" in wb.sheetnames
        assert "Generic Submissions" in wb.sheetnames

        # Check key row exists on both tabs
        for sheet_name in ["DPD + NOC + Patents", "Generic Submissions"]:
            ws = wb[sheet_name]
            row1 = [ws.cell(1, c).value for c in range(1, 5)]
            assert row1[0] == "PRODUCT KEY:", (
                f"Sheet {sheet_name!r}: row 1 must start with 'PRODUCT KEY:'"
            )

        # Same product name colors on both tabs (cell B1 fill color)
        ws1 = wb["DPD + NOC + Patents"]
        ws2 = wb["Generic Submissions"]
        fill1 = ws1.cell(1, 2).fill.fgColor.rgb if ws1.cell(1, 2).fill else None
        fill2 = ws2.cell(1, 2).fill.fgColor.rgb if ws2.cell(1, 2).fill else None
        assert fill1 == fill2, (
            f"Product 1 color must match across tabs: tab1={fill1} tab2={fill2}"
        )

    def test_four_products_palette_cycles(self):
        """Four products each get a distinct color; the key has four entries."""
        from app.enrichment.workbook import build_workbook_multiproduct, _BLOCK_COLORS
        products_data = [
            ("alpelisib",   _make_response_for("alpelisib",   ["02498014"])),
            ("apremilast",  _make_response_for("apremilast",  ["02368975"])),
            ("abrocitinib", _make_response_for("abrocitinib", ["02517019"])),
            ("pembrolizumab", _make_response_for("pembrolizumab", ["02449110"])),
        ]
        xlsx, _, _ = build_workbook_multiproduct(products_data)
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        ws = wb["DPD + NOC + Patents"]
        row1 = [ws.cell(1, c).value for c in range(1, 10)]
        product_names = [v for v in row1 if v and v != "PRODUCT KEY:"]
        assert len(product_names) == 4, f"Expected 4 products in key row, got: {product_names}"

    def test_spacer_column_between_blocks(self):
        """There is at least one empty (spacer) column between adjacent blocks."""
        from app.enrichment.workbook import build_workbook_multiproduct
        r1 = _make_response_for("alpelisib", ["02498014"])
        r2 = _make_response_for("apremilast", ["02368975"])
        xlsx, _, _ = build_workbook_multiproduct([("alpelisib", r1), ("apremilast", r2)])
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        ws = wb["DPD + NOC + Patents"]
        # In row 3 (headers), find the first None after a non-None — that's the spacer
        header_row = [ws.cell(3, c).value for c in range(1, 200) if c <= ws.max_column]
        none_positions = [i for i, v in enumerate(header_row) if v is None]
        assert none_positions, "No spacer column found between blocks in header row"

    def test_sheet2_same_products_same_colors(self):
        """Generic Submissions tab has both products with the same key as Tab 1."""
        from app.enrichment.workbook import build_workbook_multiproduct
        r1 = _make_response_for("alpelisib", ["02498014"])
        r2 = _make_response_for("apremilast", ["02368975"])
        xlsx, _, s2 = build_workbook_multiproduct([("alpelisib", r1), ("apremilast", r2)])
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=True)
        ws2 = wb["Generic Submissions"]
        row1 = [ws2.cell(1, c).value for c in range(1, 10)]
        assert row1[0] == "PRODUCT KEY:", "Generic Submissions tab must have key row"
        assert "alpelisib" in row1
        assert "apremilast" in row1
